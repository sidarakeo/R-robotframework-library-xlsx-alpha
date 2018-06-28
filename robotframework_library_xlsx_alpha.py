__author__ = 'KEO_Sidara'
__Version__ = '0.1.1'
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
import datetime
import os
from contextlib import closing

def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)
def get_row_count_xlsx(file_name):
    with closing(load_workbook(filename=file_name)) as wb:
		ws = wb.active
		row_count = ws.max_row
		return	row_count      	
def add_value_by_list(file_name, start_column, start_row, value):
    with closing(load_workbook(filename=file_name)) as wb:
   	ws = wb.active
   	start_row =int(start_row)
	start_column= int(start_column)
   	for i in value:
	    ws.cell(row=start_row, column=start_column).value = i
	    start_row += 1
  	wb.save(file_name)
def add_value_by_date(file_name, start_column, start_row, value):
    with closing(load_workbook(filename=file_name)) as wb:
	ws = wb.active
	start_row =int(start_row)
	start_column= int(start_column)
	for i in value:
		i = datetime.datetime.strptime(i, "%m/%d/%Y")
		ws.cell(row=start_row, column=start_column).value = i
		ws.cell(row=start_row, column=start_column).number_format = 'm/DD/YYYY'
		start_row += 1
	wb.save(file_name)   	    
def add_value(file_name, cell_cords, value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb.active
        ws[cell_cords] = value
        wb.save(file_name)       
def add_value_dateformat(file_name, cell_cords, value):
    with closing(load_workbook(filename=file_name)) as wb:
	ws = wb.active
	print(value)
	valuec = datetime.datetime.strptime(value, "%m/%d/%Y")
	cell = ws[cell_cords]
	cell.value = valuec
	cell.number_format = 'm/DD/YYYY'
	wb.save(file_name)
		
def convert_date_format(value):
		value = datetime.datetime.strptime(value, "%m/%d/%Y").strftime('%m/%d/%Y').lstrip('0')
		#value.strftime("%m/%d/%Y").lstrip('0')
		return value
		#print(value)			
def add_value_generalformat(file_name, cell_cords, value):
    with closing(load_workbook(filename=file_name)) as wb:
		ws = wb.active
		cell = ws[cell_cords]
		cell.value = value
		cell.number_format = 'General'
		wb.save(file_name)
def add_value_number_format(file_name, cell_cords, value):
    with closing(load_workbook(filename=file_name)) as wb:
		ws = wb.active
		cell = ws[cell_cords]
		cell.value = value
		cell.number_format = 'Number'
		wb.save(file_name)			
def run_bat(file_name, filedir):
		owd = os.getcwd()
		filepath= file_name
		os.chdir(filedir)
		os.startfile(file_name)
		os.chdir(owd)

